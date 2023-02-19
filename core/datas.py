'''
@Filename: automatedTesting -> datas
@Author: AnnieZhang
@Date: 2023/1/28 9:41
'''
from csv import DictReader
from openpyxl import load_workbook

def read_csv(file):
    with open(file, encoding="utf-8-sig") as f:
        reader = DictReader(f) # 实例化csv读取器
        for d in reader:
            yield d # 带有yield的函数是一个生成器函数，返回的是迭代器  https://www.py.cn/faq/python/21543.html

def read_excel(file):
    wb = load_workbook(file) # 打开文件
    ws = wb.active
    for d in ws.iter_rows(values_only=True, min_row=2):
        yield d


if __name__ == "__main__":
    read = read_excel(file=r"../datas/annie.xlsx")
    for r in read:
        print(r)
        print(type(r))

